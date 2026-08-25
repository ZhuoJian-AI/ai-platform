"""Focused tests for direct Open Agent Skill script execution."""

from __future__ import annotations

import asyncio
import base64
import hashlib
import io
import json
import os
import time
import zipfile
from pathlib import Path

import app as runner
import builtin_tools as builtin
import httpx
import pytest
from PIL import Image


@pytest.mark.asyncio
async def test_runner_capacity_queues_and_releases_slots():
    capacity = runner.RunnerCapacity(limit=1, queue_limit=2, wait_seconds=2, label="test")
    first_entered = asyncio.Event()
    release_first = asyncio.Event()

    async def first():
        async with capacity.slot():
            first_entered.set()
            await release_first.wait()

    async def second():
        async with capacity.slot():
            return "second"

    first_task = asyncio.create_task(first())
    await first_entered.wait()
    second_task = asyncio.create_task(second())
    await asyncio.sleep(0)
    assert capacity.snapshot() == {"active": 1, "waiting": 1, "limit": 1}
    release_first.set()
    await first_task
    assert await second_task == "second"
    assert capacity.snapshot() == {"active": 0, "waiting": 0, "limit": 1}


@pytest.mark.asyncio
async def test_runner_capacity_rejects_full_queue():
    capacity = runner.RunnerCapacity(limit=1, queue_limit=1, wait_seconds=2, label="test")
    first_entered = asyncio.Event()
    release_first = asyncio.Event()

    async def hold():
        async with capacity.slot():
            first_entered.set()
            await release_first.wait()

    holder = asyncio.create_task(hold())
    await first_entered.wait()
    waiting = asyncio.create_task(capacity._wait_for_slot())
    await asyncio.sleep(0)
    with pytest.raises(runner.HTTPException) as exc:
        await capacity._wait_for_slot()
    assert exc.value.status_code == 429
    waiting.cancel()
    await asyncio.gather(waiting, return_exceptions=True)
    release_first.set()
    await holder


def _package(script_path: str, content: bytes) -> tuple[str, str]:
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("SKILL.md", "---\nname: Test\ndescription: Test script\n---\n")
        zf.writestr(script_path, content)
    raw = out.getvalue()
    return hashlib.sha256(raw).hexdigest(), base64.b64encode(raw).decode("ascii")


def _package_files(files: dict[str, bytes]) -> tuple[str, str]:
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("SKILL.md", "---\nname: Test\ndescription: Test script\n---\n")
        for path, content in files.items():
            zf.writestr(path, content)
    raw = out.getvalue()
    return hashlib.sha256(raw).hexdigest(), base64.b64encode(raw).decode("ascii")


@pytest.mark.asyncio
async def test_signed_archive_download_verifies_size_and_checksum(monkeypatch):
    package_hash, encoded = _package("scripts/run.py", b"print('ok')")
    raw = base64.b64decode(encoded)

    def handler(request: httpx.Request) -> httpx.Response:
        assert request.url == httpx.URL("https://storage.invalid/package.zip")
        assert request.headers["x-download-token"] == "short-lived"
        return httpx.Response(200, content=raw)

    transport = httpx.MockTransport(handler)
    real_client = httpx.AsyncClient

    def client_factory(*_args, **_kwargs):
        return real_client(transport=transport)

    monkeypatch.setattr(runner.httpx, "AsyncClient", client_factory)
    request = runner.InstallRequest(
        package_hash=package_hash,
        archive_url="https://storage.invalid/package.zip",
        archive_headers={"X-Download-Token": "short-lived"},
        archive_size=len(raw),
        runtime="agent_skill",
    )
    assert await runner._load_archive(request) == raw


def test_skill_archive_rejects_duplicate_and_unsafe_paths(tmp_path):
    duplicate = io.BytesIO()
    with zipfile.ZipFile(duplicate, "w") as archive:
        archive.writestr("SKILL.md", "one")
        archive.writestr("skill.md", "two")
    with pytest.raises(runner.HTTPException, match="Duplicate archive path"):
        runner._extract(duplicate.getvalue(), tmp_path / "duplicate")

    unsafe = io.BytesIO()
    with zipfile.ZipFile(unsafe, "w") as archive:
        archive.writestr("../outside.py", "blocked")
    with pytest.raises(runner.HTTPException, match="Unsafe archive path"):
        runner._extract(unsafe.getvalue(), tmp_path / "unsafe")


def test_skill_archive_enforces_file_count_and_expanded_size(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "MAX_PACKAGE_FILES", 1)
    too_many = io.BytesIO()
    with zipfile.ZipFile(too_many, "w") as archive:
        archive.writestr("SKILL.md", "manifest")
        archive.writestr("scripts/run.py", "print('x')")
    with pytest.raises(runner.HTTPException, match="more than 1 files"):
        runner._extract(too_many.getvalue(), tmp_path / "many")

    monkeypatch.setattr(runner, "MAX_PACKAGE_FILES", 1000)
    monkeypatch.setattr(runner, "MAX_EXPANDED_PACKAGE_BYTES", 4)
    expanded = io.BytesIO()
    with zipfile.ZipFile(expanded, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("SKILL.md", "12345")
    with pytest.raises(runner.HTTPException, match="Expanded Skill package"):
        runner._extract(expanded.getvalue(), tmp_path / "expanded")


@pytest.mark.asyncio
async def test_python_agent_skill_uses_direct_script_and_io_directories(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    script = b"""import os
from pathlib import Path
source = next(Path(os.environ['SKILL_INPUT_DIR']).iterdir())
target = Path(os.environ['SKILL_OUTPUT_DIR']) / 'cleaned.txt'
target.write_text(source.read_text(encoding='utf-8').upper(), encoding='utf-8')
"""
    package_hash, archive = _package("scripts/clean.py", script)
    request = runner.ExecuteRequest(
        package_hash=package_hash,
        archive_base64=archive,
        runtime="agent_skill",
        script_path="scripts/clean.py",
        inputs=[runner.InputFile(
            file_id="1", name="input.txt",
            content_base64=base64.b64encode(b"hello").decode(),
        )],
        execution_id=1,
    )
    result = await runner.execute(request, runner.RUNNER_TOKEN)
    assert result["status"] == "success"
    assert result["outputs"][0]["name"] == "cleaned.txt"
    assert base64.b64decode(result["outputs"][0]["content_base64"]).decode() == "HELLO"


@pytest.mark.asyncio
async def test_skill_process_does_not_inherit_platform_secrets(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    monkeypatch.setenv("DATABASE_URL", "postgresql://must-not-leak")
    monkeypatch.setenv("REDIS_URL", "redis://must-not-leak")
    monkeypatch.setenv("MODEL_API_KEY", "must-not-leak")
    script = b"""import json
import os
from pathlib import Path

names = ["DATABASE_URL", "REDIS_URL", "MODEL_API_KEY"]
Path(os.environ["SKILL_OUTPUT_DIR"], "environment.json").write_text(
    json.dumps({name: os.environ.get(name) for name in names}), encoding="utf-8"
)
"""
    package_hash, archive = _package("scripts/environment.py", script)
    result = await runner.execute(
        runner.ExecuteRequest(
            package_hash=package_hash,
            archive_base64=archive,
            runtime="agent_skill",
            script_path="scripts/environment.py",
            execution_id=11,
        ),
        runner.RUNNER_TOKEN,
    )
    output = next(item for item in result["outputs"] if item["name"] == "environment.json")
    assert json.loads(base64.b64decode(output["content_base64"])) == {
        "DATABASE_URL": None,
        "REDIS_URL": None,
        "MODEL_API_KEY": None,
    }


@pytest.mark.asyncio
async def test_execution_temp_directory_is_removed(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    created: list[Path] = []
    real_mkdtemp = runner.tempfile.mkdtemp

    def tracked_mkdtemp(*args, **kwargs):
        path = Path(real_mkdtemp(*args, **kwargs))
        if str(kwargs.get("prefix", "")).startswith("run-"):
            created.append(path)
        return str(path)

    monkeypatch.setattr(runner.tempfile, "mkdtemp", tracked_mkdtemp)
    package_hash, archive = _package("scripts/run.py", b"print('done')")
    await runner.execute(
        runner.ExecuteRequest(
            package_hash=package_hash,
            archive_base64=archive,
            runtime="agent_skill",
            script_path="scripts/run.py",
            execution_id=12,
        ),
        runner.RUNNER_TOKEN,
    )
    assert len(created) == 1
    assert not created[0].exists()


@pytest.mark.asyncio
async def test_runner_cache_removes_expired_failed_and_lru_entries(tmp_path, monkeypatch):
    cache = tmp_path / "cache"
    cache.mkdir()
    monkeypatch.setattr(runner, "CACHE_ROOT", cache)
    monkeypatch.setattr(runner, "CACHE_RETENTION_SECONDS", 60)
    monkeypatch.setattr(runner, "FAILED_TEMP_RETENTION_SECONDS", 30)
    monkeypatch.setattr(runner, "CACHE_MAX_BYTES", 12)
    runner._ACTIVE_PACKAGE_HASHES.clear()

    old_hash = "a" * 64
    old = cache / old_hash
    old.mkdir()
    (old / "payload").write_bytes(b"old")
    (old / ".last_used").touch()
    os.utime(old / ".last_used", (time.time() - 120, time.time() - 120))

    failed = cache / "install-failed"
    failed.mkdir()
    (failed / "payload").write_bytes(b"failed")
    os.utime(failed, (time.time() - 60, time.time() - 60))

    least_recent_hash = "b" * 64
    least_recent = cache / least_recent_hash
    least_recent.mkdir()
    (least_recent / "payload").write_bytes(b"12345678")
    (least_recent / ".last_used").touch()
    os.utime(least_recent / ".last_used", (time.time() - 20, time.time() - 20))

    recent_hash = "c" * 64
    recent = cache / recent_hash
    recent.mkdir()
    (recent / "payload").write_bytes(b"abcdefgh")
    (recent / ".last_used").touch()

    active_hash = "d" * 64
    active = cache / active_hash
    active.mkdir()
    (active / "payload").write_bytes(b"active-cache")
    (active / ".last_used").touch()
    runner._ACTIVE_PACKAGE_HASHES.add(active_hash)
    try:
        result = await runner._cleanup_cache()
    finally:
        runner._ACTIVE_PACKAGE_HASHES.discard(active_hash)

    assert not old.exists()
    assert not failed.exists()
    assert not least_recent.exists()
    assert not recent.exists()
    assert active.exists()
    assert result["removed_entries"] == 4
    assert result["active_entries"] == 1


@pytest.mark.asyncio
async def test_standard_script_must_stay_under_scripts(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    package_hash, archive = _package("scripts/clean.py", b"print('ok')")
    request = runner.ExecuteRequest(
        package_hash=package_hash,
        archive_base64=archive,
        runtime="agent_skill",
        script_path="../scripts/clean.py",
        execution_id=2,
    )
    with pytest.raises(runner.HTTPException) as exc:
        await runner.execute(request, runner.RUNNER_TOKEN)
    assert exc.value.status_code == 422


@pytest.mark.parametrize("script_path,language", [
    ("scripts/a.py", "python"),
    ("scripts/a.js", "node"),
    ("scripts/a.sh", "bash"),
])
def test_supported_script_languages_are_detected(tmp_path: Path, script_path: str, language: str):
    package = tmp_path / "package"
    target = package / script_path
    target.parent.mkdir(parents=True)
    target.write_text("", encoding="utf-8")
    _, actual = runner._script(package.resolve(), script_path, None)
    assert actual == language


def test_python_version_requirement_rejects_incompatible_skill(tmp_path: Path, monkeypatch):
    (tmp_path / "scripts").mkdir()
    (tmp_path / "scripts" / "run.py").write_text("print('ok')", encoding="utf-8")
    (tmp_path / "pyproject.toml").write_text(
        '[project]\nname="future-skill"\nversion="1.0.0"\nrequires-python=">=99"\n', encoding="utf-8",
    )
    monkeypatch.setattr(runner, "_runtime_info", lambda: {
        "python_version": "3.12.0", "node_version": "20.0.0",
        "bash_version": None, "libreoffice_version": None, "builtin_dependencies": {},
    })
    with pytest.raises(runner.HTTPException) as exc:
        runner._validate_runtime_compatibility(tmp_path)
    assert "requires Python >=99" in exc.value.detail


def test_node_engine_requirement_and_missing_version_warning(tmp_path: Path, monkeypatch):
    (tmp_path / "scripts").mkdir()
    (tmp_path / "scripts" / "run.js").write_text("console.log('ok')", encoding="utf-8")
    (tmp_path / "package.json").write_text('{"engines":{"node":">=20 <21"}}', encoding="utf-8")
    monkeypatch.setattr(runner, "_runtime_info", lambda: {
        "python_version": "3.12.0", "node_version": "20.20.2",
        "bash_version": None, "libreoffice_version": None, "builtin_dependencies": {},
    })
    assert runner._validate_runtime_compatibility(tmp_path) == []
    (tmp_path / "package.json").write_text('{}', encoding="utf-8")
    warnings = runner._validate_runtime_compatibility(tmp_path)
    assert any("engines.node" in warning for warning in warnings)


@pytest.mark.asyncio
async def test_pyproject_takes_precedence_over_requirements(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    calls: list[list[str]] = []

    async def fake_run(argv, *_args, **_kwargs):
        calls.append(argv)
        return 0, "", ""

    monkeypatch.setattr(runner, "_run", fake_run)
    monkeypatch.setattr(runner, "_runtime_info", lambda: {
        "python_version": "3.12.0", "node_version": "20.0.0",
        "bash_version": None, "libreoffice_version": None, "builtin_dependencies": {},
    })
    package_hash, archive = _package_files({
        "scripts/run.py": b"print('ok')",
        "pyproject.toml": b'[project]\nname="priority-test"\nversion="1.0.0"\nrequires-python=">=3.12"\n',
        "requirements.txt": b"this-must-not-be-installed==999\n",
    })
    await runner._ensure_installed(runner.InstallRequest(
        package_hash=package_hash, archive_base64=archive, runtime="agent_skill",
    ))
    pip_calls = [call for call in calls if "pip" in call]
    assert any(call[-2:] == ["install", "."] for call in pip_calls)
    assert not any("-r" in call for call in pip_calls)


@pytest.mark.asyncio
@pytest.mark.parametrize(("script_path", "content", "output_name", "expected"), [
    (
        "scripts/write.js",
        b"require('fs').writeFileSync(process.env.SKILL_OUTPUT_DIR + '/node.txt', 'NODE')\n",
        "node.txt",
        b"NODE",
    ),
    (
        "scripts/write.sh",
        b"printf BASH > \"$SKILL_OUTPUT_DIR/bash.txt\"\n",
        "bash.txt",
        b"BASH",
    ),
])
async def test_node_and_bash_scripts_execute_directly(
    tmp_path, monkeypatch, script_path: str, content: bytes, output_name: str, expected: bytes,
):
    if script_path.endswith(".sh") and runner._command_version(["bash", "--version"]) is None:
        pytest.skip("A functional Bash interpreter is not available on this host")
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    package_hash, archive = _package(script_path, content)
    result = await runner.execute(runner.ExecuteRequest(
        package_hash=package_hash,
        archive_base64=archive,
        runtime="agent_skill",
        script_path=script_path,
        execution_id=3,
    ), runner.RUNNER_TOKEN)
    output = next(item for item in result["outputs"] if item["name"] == output_name)
    assert base64.b64decode(output["content_base64"]) == expected


@pytest.mark.asyncio
async def test_python_skill_without_requirements_uses_builtin_excel_library(
    tmp_path, monkeypatch,
):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    script = b"""from pathlib import Path
import os
from openpyxl import Workbook

book = Workbook()
book.active.append(["name", "value"])
book.active.append(["ready", 1])
book.save(Path(os.environ["SKILL_OUTPUT_DIR"]) / "result.xlsx")
"""
    package_hash, archive = _package("scripts/create_excel.py", script)
    result = await runner.execute(
        runner.ExecuteRequest(
            package_hash=package_hash,
            archive_base64=archive,
            runtime="agent_skill",
            script_path="scripts/create_excel.py",
            execution_id=4,
        ),
        runner.RUNNER_TOKEN,
    )
    output = next(item for item in result["outputs"] if item["name"] == "result.xlsx")
    assert base64.b64decode(output["content_base64"]).startswith(b"PK")


@pytest.mark.asyncio
async def test_node_skill_without_package_json_uses_builtin_exceljs(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    script = b"""const ExcelJS = require('exceljs');
const path = require('path');
const workbook = new ExcelJS.Workbook();
workbook.addWorksheet('Sheet1').addRow(['ready', 1]);
(async () => {
  await workbook.xlsx.writeFile(path.join(process.env.SKILL_OUTPUT_DIR, 'result.xlsx'));
})();
"""
    package_hash, archive = _package("scripts/create_excel.js", script)
    result = await runner.execute(
        runner.ExecuteRequest(
            package_hash=package_hash,
            archive_base64=archive,
            runtime="agent_skill",
            script_path="scripts/create_excel.js",
            execution_id=5,
        ),
        runner.RUNNER_TOKEN,
    )
    output = next(item for item in result["outputs"] if item["name"] == "result.xlsx")
    assert base64.b64decode(output["content_base64"]).startswith(b"PK")


@pytest.mark.asyncio
async def test_python_skill_installs_extra_dependency_in_its_own_environment(
    tmp_path, monkeypatch,
):
    monkeypatch.setattr(runner, "CACHE_ROOT", tmp_path / "cache")
    package_hash, archive = _package_files({
        "requirements.txt": b"humanize==4.12.1\n",
        "scripts/run.py": b"""from pathlib import Path
import os
import humanize

Path(os.environ["SKILL_OUTPUT_DIR"], "result.txt").write_text(
    humanize.intcomma(1234567), encoding="utf-8"
)
""",
    })
    result = await runner.execute(
        runner.ExecuteRequest(
            package_hash=package_hash,
            archive_base64=archive,
            runtime="agent_skill",
            script_path="scripts/run.py",
            execution_id=6,
        ),
        runner.RUNNER_TOKEN,
    )
    output = next(item for item in result["outputs"] if item["name"] == "result.txt")
    assert base64.b64decode(output["content_base64"]) == b"1,234,567"
    metadata = json.loads((tmp_path / "cache" / package_hash / ".install.json").read_text())
    assert metadata["installed_dependencies"]["python"] == ["humanize==4.12.1"]


@pytest.mark.asyncio
async def test_builtin_spreadsheet_create_does_not_load_skill_package(monkeypatch):
    async def fail_install(*_args, **_kwargs):
        raise AssertionError("builtin execution must not install or load a Skill package")

    monkeypatch.setattr(runner, "_ensure_installed", fail_install)
    result = await runner.execute_builtin_tool(
        runner.BuiltinExecuteRequest(
            tool_kind="spreadsheet",
            action="create",
            params={
                "output_name": "测试表.xlsx",
                "sheets": [{"name": "数据", "rows": [["名称", "金额"], ["样例", 123.45]]}],
            },
            execution_id="builtin-create",
        ),
        runner.RUNNER_TOKEN,
    )
    assert result["status"] == "success"
    assert result["tool_kind"] == "spreadsheet"
    output = result["outputs"][0]
    assert output["name"] == "测试表.xlsx"
    assert base64.b64decode(output["content_base64"]).startswith(b"PK")


@pytest.mark.asyncio
async def test_builtin_spreadsheet_inspects_original_input():
    from openpyxl import Workbook

    stream = io.BytesIO()
    book = Workbook()
    book.active.title = "明细"
    book.active.append(["姓名", "金额"])
    book.active.append(["张三", 88])
    book.save(stream)
    result = await runner.execute_builtin_tool(
        runner.BuiltinExecuteRequest(
            tool_kind="spreadsheet",
            action="inspect",
            inputs=[runner.InputFile(
                file_id="file-1",
                name="明细.xlsx",
                content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
            )],
            execution_id="builtin-inspect",
        ),
        runner.RUNNER_TOKEN,
    )
    assert result["outputs"] == []
    assert result["summary"]["sheets"][0]["name"] == "明细"
    assert result["summary"]["sheets"][0]["rows"][1] == ["张三", 88]


@pytest.mark.asyncio
async def test_builtin_execution_requires_internal_token():
    with pytest.raises(runner.HTTPException) as exc:
        await runner.execute_builtin_tool(
            runner.BuiltinExecuteRequest(
                tool_kind="text", action="create", params={"content": "x"}, execution_id="no-auth",
            ),
            None,
        )
    assert exc.value.status_code == 401


@pytest.mark.asyncio
async def test_builtin_web_fetch_extracts_readable_html(monkeypatch):
    monkeypatch.setattr(
        builtin,
        "_http_get",
        lambda *_args, **_kwargs: (
            "https://example.com/article",
            "<html><head><title>测试文章</title><script>bad()</script></head>"
            "<body><h1>标题</h1><p>正文内容</p></body></html>".encode(),
            "text/html",
            "",
        ),
    )
    result = await runner.execute_builtin_tool(
        runner.BuiltinExecuteRequest(
            tool_kind="web",
            action="fetch",
            params={"url": "https://example.com/article"},
            execution_id="web-fetch",
        ),
        runner.RUNNER_TOKEN,
    )
    assert result["summary"]["title"] == "测试文章"
    assert "正文内容" in result["summary"]["content"]
    assert "bad()" not in result["summary"]["content"]


@pytest.mark.asyncio
async def test_builtin_web_search_falls_back_to_bing(monkeypatch):
    def fake_http_get(url, **_kwargs):
        if "duckduckgo" in url:
            return url, b"<html><body>challenge</body></html>", "text/html", ""
        html = (
            b'<html><body><ol><li class="b_algo"><h2><a href="https://example.com">'
            b'Example result</a></h2><div class="b_caption"><p>Useful snippet</p></div>'
            b"</li></ol></body></html>"
        )
        return (
            url,
            html,
            "text/html",
            "",
        )

    monkeypatch.setattr(builtin, "_http_get", fake_http_get)
    result = await runner.execute_builtin_tool(
        runner.BuiltinExecuteRequest(
            tool_kind="web",
            action="search",
            params={"query": "example", "max_results": 3},
            execution_id="web-search",
        ),
        runner.RUNNER_TOKEN,
    )
    assert result["summary"]["results"] == [{
        "title": "Example result",
        "url": "https://example.com",
        "snippet": "Useful snippet",
    }]


def test_builtin_web_rejects_private_and_invalid_urls():
    with pytest.raises(builtin.BuiltinToolError, match="private"):
        builtin._validate_public_url("http://127.0.0.1/secret")
    with pytest.raises(builtin.BuiltinToolError, match="port"):
        builtin._validate_public_url("https://example.com:not-a-port/")


@pytest.mark.asyncio
async def test_builtin_image_resize_outputs_real_png():
    stream = io.BytesIO()
    Image.new("RGB", (120, 60), "red").save(stream, format="PNG")
    result = await runner.execute_builtin_tool(
        runner.BuiltinExecuteRequest(
            tool_kind="image",
            action="resize",
            params={"width": 30, "height": 30, "output_name": "缩略图.png"},
            inputs=[runner.InputFile(
                file_id="image-1",
                name="原图.png",
                content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
            )],
            execution_id="image-resize",
        ),
        runner.RUNNER_TOKEN,
    )
    output = result["outputs"][0]
    with Image.open(io.BytesIO(base64.b64decode(output["content_base64"]))) as image:
        assert image.size == (30, 15)


@pytest.mark.asyncio
async def test_builtin_archive_extract_preserves_relative_paths():
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("资料/说明.txt", "内容")
        archive.writestr("数据.csv", "a,b\n1,2")
    result = await runner.execute_builtin_tool(
        runner.BuiltinExecuteRequest(
            tool_kind="archive",
            action="extract",
            inputs=[runner.InputFile(
                file_id="archive-1",
                name="资料包.zip",
                content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
            )],
            execution_id="archive-extract",
        ),
        runner.RUNNER_TOKEN,
    )
    assert {item["relative_path"] for item in result["outputs"]} == {"资料/说明.txt", "数据.csv"}


@pytest.mark.asyncio
async def test_builtin_archive_rejects_path_traversal():
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as archive:
        archive.writestr("../escape.txt", "blocked")
    with pytest.raises(runner.HTTPException) as exc:
        await runner.execute_builtin_tool(
            runner.BuiltinExecuteRequest(
                tool_kind="archive",
                action="extract",
                inputs=[runner.InputFile(
                    file_id="archive-2",
                    name="unsafe.zip",
                    content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
                )],
                execution_id="archive-unsafe",
            ),
            runner.RUNNER_TOKEN,
        )
    assert exc.value.status_code == 422
    assert "Unsafe archive path" in str(exc.value.detail)


@pytest.mark.asyncio
async def test_builtin_archive_rejects_zip_links():
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as archive:
        link = zipfile.ZipInfo("link.txt")
        link.create_system = 3
        link.external_attr = 0o120777 << 16
        archive.writestr(link, "target.txt")
    with pytest.raises(runner.HTTPException) as exc:
        await runner.execute_builtin_tool(
            runner.BuiltinExecuteRequest(
                tool_kind="archive",
                action="extract",
                inputs=[runner.InputFile(
                    file_id="archive-link",
                    name="unsafe-link.zip",
                    content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
                )],
                execution_id="archive-link",
            ),
            runner.RUNNER_TOKEN,
        )
    assert exc.value.status_code == 422
    assert "links" in str(exc.value.detail).lower()
