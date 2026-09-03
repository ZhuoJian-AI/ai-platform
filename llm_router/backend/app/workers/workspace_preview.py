"""Single-concurrency, crash-recoverable LibreOffice fallback preview worker."""

from __future__ import annotations

import asyncio
import csv
import gzip
import json
import os
import shutil
import socket
import tempfile
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from pathlib import Path, PurePosixPath

from sqlalchemy import or_, select, text

from app.config import settings
from app.database import async_session_factory, engine
from app.models.workspace import WorkspaceFileVersion, WorkspacePreviewJob
from app.services import storage_gateway_service

WORKER_ID = f"{socket.gethostname()}:{os.getpid()}"
CONVERSION_TIMEOUT_SECONDS = 10 * 60
MAX_OUTPUT_BYTES = 500 * 1024 * 1024
MAX_SPREADSHEET_OUTPUT_BYTES = 100 * 1024 * 1024
SPREADSHEET_MAX_ROWS = 10_000
SPREADSHEET_MAX_COLUMNS = 100
SPREADSHEET_PAGE_SIZE = 200
SPREADSHEET_TEXT_BUDGET = 10_000_000
WORKER_ADVISORY_LOCK_ID = 9_318_411


async def claim_one() -> str | None:
    now = datetime.now(UTC)
    async with async_session_factory() as db:
        async with db.begin():
            row = (await db.execute(
                select(WorkspacePreviewJob)
                .where(
                    WorkspacePreviewJob.attempt_count < 3,
                    WorkspacePreviewJob.next_attempt_at <= now,
                    or_(
                        WorkspacePreviewJob.status == "queued",
                        (WorkspacePreviewJob.status == "processing")
                        & (WorkspacePreviewJob.lease_expires_at < now),
                    ),
                )
                .order_by(WorkspacePreviewJob.next_attempt_at, WorkspacePreviewJob.created_at)
                .with_for_update(skip_locked=True)
                .limit(1)
            )).scalar_one_or_none()
            if row is None:
                return None
            row.status = "processing"
            row.attempt_count = int(row.attempt_count or 0) + 1
            row.locked_by = WORKER_ID
            row.lease_expires_at = now + timedelta(
                seconds=settings.workspace_preview_job_lease_seconds
            )
            row.error = None
            return str(row.id)


async def _convert_to_pdf(source: Path, output_dir: Path) -> Path:
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        raise RuntimeError("LibreOffice unavailable")
    profile = output_dir / "lo-profile"
    profile.mkdir()
    process = await asyncio.create_subprocess_exec(
        executable,
        "--headless",
        f"-env:UserInstallation={profile.as_uri()}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_dir),
        str(source),
        stdout=asyncio.subprocess.DEVNULL,
        stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        await asyncio.wait_for(process.wait(), timeout=CONVERSION_TIMEOUT_SECONDS)
    except TimeoutError:
        process.kill()
        await process.wait()
        raise RuntimeError("LibreOffice timeout")
    if process.returncode:
        raise RuntimeError("LibreOffice conversion failed")
    output = output_dir / f"{source.stem}.pdf"
    if not output.is_file() or output.stat().st_size <= 0:
        raise RuntimeError("LibreOffice produced no PDF")
    return output


async def _convert_to_xlsx(source: Path, output_dir: Path) -> Path:
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        raise RuntimeError("LibreOffice unavailable")
    profile = output_dir / "lo-xlsx-profile"
    profile.mkdir()
    process = await asyncio.create_subprocess_exec(
        executable, "--headless", f"-env:UserInstallation={profile.as_uri()}",
        "--convert-to", "xlsx", "--outdir", str(output_dir), str(source),
        stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        await asyncio.wait_for(process.wait(), timeout=CONVERSION_TIMEOUT_SECONDS)
    except TimeoutError:
        process.kill()
        await process.wait()
        raise RuntimeError("LibreOffice timeout")
    output = output_dir / f"{source.stem}.xlsx"
    if process.returncode or not output.is_file() or output.stat().st_size <= 0:
        raise RuntimeError("LibreOffice spreadsheet conversion failed")
    return output


def _cell_value(value: object) -> str | int | float | bool | None:
    if isinstance(value, str):
        return value[:512]
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)[:512]


async def _spreadsheet_rows(source: Path, output_dir: Path) -> Path:
    from openpyxl import load_workbook

    if source.suffix.lower() in {".csv", ".tsv"}:
        def build_delimited() -> dict:
            rows: list[list[str | int | float | bool | None]] = []
            used = 0
            truncated = False
            with source.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                reader = csv.reader(handle, delimiter="\t" if source.suffix.lower() == ".tsv" else ",")
                for row_number, row in enumerate(reader, start=1):
                    if row_number > SPREADSHEET_MAX_ROWS:
                        truncated = True
                        break
                    values = [_cell_value(value) for value in row[:SPREADSHEET_MAX_COLUMNS]]
                    used += sum(len(str(value)) for value in values if value is not None)
                    if used > SPREADSHEET_TEXT_BUDGET:
                        truncated = True
                        break
                    rows.append(values)
            return {
                "page_size": SPREADSHEET_PAGE_SIZE,
                "sheets": [{
                    "name": "CSV" if source.suffix.lower() == ".csv" else "TSV",
                    "total_rows": len(rows),
                    "columns": max((len(row) for row in rows), default=0),
                    "truncated": truncated,
                    "pages": [
                        rows[index:index + SPREADSHEET_PAGE_SIZE]
                        for index in range(0, len(rows), SPREADSHEET_PAGE_SIZE)
                    ],
                }],
            }

        payload = await asyncio.to_thread(build_delimited)
    else:
        workbook_source = source
        if source.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            workbook_source = await _convert_to_xlsx(source, output_dir)

        def build_workbook() -> dict:
            workbook = load_workbook(workbook_source, read_only=True, data_only=True, keep_links=False)
            sheets: list[dict] = []
            used = 0
            try:
                for worksheet in workbook.worksheets:
                    rows: list[list[str | int | float | bool | None]] = []
                    truncated = False
                    for row_number, row in enumerate(
                        worksheet.iter_rows(max_col=SPREADSHEET_MAX_COLUMNS, values_only=True), start=1,
                    ):
                        if row_number > SPREADSHEET_MAX_ROWS:
                            truncated = True
                            break
                        values = [_cell_value(value) for value in row]
                        used += sum(len(str(value)) for value in values if value is not None)
                        if used > SPREADSHEET_TEXT_BUDGET:
                            truncated = True
                            break
                        while values and values[-1] is None:
                            values.pop()
                        rows.append(values)
                    sheets.append({
                        "name": str(worksheet.title),
                        "total_rows": len(rows),
                        "columns": max((len(row) for row in rows), default=0),
                        "truncated": truncated,
                        "pages": [
                            rows[index:index + SPREADSHEET_PAGE_SIZE]
                            for index in range(0, len(rows), SPREADSHEET_PAGE_SIZE)
                        ],
                    })
            finally:
                workbook.close()
            return {"page_size": SPREADSHEET_PAGE_SIZE, "sheets": sheets}

        payload = await asyncio.to_thread(build_workbook)

    output = output_dir / "spreadsheet-preview.json.gz"
    with gzip.open(output, "wt", encoding="utf-8", compresslevel=6) as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
    if output.stat().st_size > MAX_SPREADSHEET_OUTPUT_BYTES:
        raise RuntimeError("Spreadsheet preview artifact exceeds limit")
    return output


async def process_one(job_id: str) -> None:
    async with async_session_factory() as db:
        job = await db.get(WorkspacePreviewJob, job_id)
        if job is None or job.status != "processing" or job.locked_by != WORKER_ID:
            return
        try:
            version = await db.get(WorkspaceFileVersion, job.file_version_id)
            if version is None:
                raise RuntimeError("Source version missing")
            if not storage_gateway_service.is_object_ref(version.content_ref):
                raise RuntimeError("Legacy source is not in object storage")
            metadata = dict(version.metadata_ or {})
            filename = str(metadata.get("name") or "document")
            suffix = PurePosixPath(filename).suffix.lower()
            if not suffix:
                raise RuntimeError("Source format is missing")
            with tempfile.TemporaryDirectory(prefix="workspace-preview-") as temp_dir:
                directory = Path(temp_dir)
                source = directory / f"source{suffix}"
                await storage_gateway_service.download_to_path(
                    str(version.content_ref),
                    source,
                    max_bytes=settings.workspace_weboffice_max_bytes,
                )
                if job.conversion_type == "spreadsheet_rows":
                    output = await _spreadsheet_rows(source, directory)
                    output_filename = f"workspace-previews/{job.file_version_id}-spreadsheet.json.gz"
                    output_type = "application/gzip"
                    output_limit = MAX_SPREADSHEET_OUTPUT_BYTES
                elif job.conversion_type == "pdf":
                    if suffix == ".pdf":
                        raise RuntimeError("Source format does not require Office conversion")
                    output = await _convert_to_pdf(source, directory)
                    output_filename = f"workspace-previews/{job.file_version_id}.pdf"
                    output_type = "application/pdf"
                    output_limit = MAX_OUTPUT_BYTES
                else:
                    raise RuntimeError("Unsupported preview conversion type")
                output_ref = await storage_gateway_service.upload_path(
                    output,
                    filename=output_filename,
                    content_type=output_type,
                    max_bytes=output_limit,
                )
            if job.output_ref and job.output_ref != output_ref:
                try:
                    await storage_gateway_service.delete_object(job.output_ref)
                except storage_gateway_service.StorageGatewayError:
                    pass
            job.output_ref = output_ref
            job.status = "ready"
            job.error = None
        except Exception as exc:
            if int(job.attempt_count or 0) >= 3:
                job.status = "failed"
                job.error = f"预览生成失败：{type(exc).__name__}"[:500]
            else:
                job.status = "queued"
                job.next_attempt_at = datetime.now(UTC) + timedelta(
                    seconds=30 * (2 ** max(int(job.attempt_count or 1) - 1, 0))
                )
                job.error = None
        finally:
            job.lease_expires_at = None
            job.locked_by = None
            await db.commit()


async def main() -> None:
    while True:
        # A database advisory lock keeps conversion globally single-concurrency,
        # even if the deployment is accidentally started with multiple replicas.
        async with engine.connect() as lock_connection:
            acquired = bool((await lock_connection.execute(
                text("SELECT pg_try_advisory_lock(:lock_id)"),
                {"lock_id": WORKER_ADVISORY_LOCK_ID},
            )).scalar_one())
            if acquired:
                try:
                    job_id = await claim_one()
                    if job_id:
                        await process_one(job_id)
                        continue
                finally:
                    await lock_connection.execute(
                        text("SELECT pg_advisory_unlock(:lock_id)"),
                        {"lock_id": WORKER_ADVISORY_LOCK_ID},
                    )
        await asyncio.sleep(settings.workspace_preview_job_poll_seconds)


if __name__ == "__main__":
    asyncio.run(main())
